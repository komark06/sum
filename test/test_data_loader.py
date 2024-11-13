from openpyxl import Workbook
from src.data_loader import ExcelDataLoader, Summons
from datetime import date
from itertools import chain
from io import BytesIO


def test_load():
    """Verify that ExcelDataLoader successfully loads data from an file.

    Create a virtual excel and use ExcelDataLoader to load.
    """
    targets = [
        Summons(
            account="20240422-5259-000069",
            date=date(2024, 4, 22),
            amount=10,
        )
    ]
    numbers = [
        Summons(
            account="20240411-5256-000107",
            date=date(2024, 4, 11),
            amount=3,
        ),
        Summons(
            account="20240411-5257-000014",
            date=date(2024, 4, 11),
            amount=4,
        ),
        Summons(
            account="20240411-5256-000094",
            date=date(2024, 4, 11),
            amount=5,
        ),
        Summons(
            account="20240411-5257-000101",
            date=date(2024, 4, 11),
            amount=5,
        ),
    ]
    workbook = Workbook()
    sheet = workbook.active
    for i, target in enumerate(targets):
        sheet.cell(i + 1, 1, target.amount)
    for i, number in enumerate(numbers):
        sheet.cell(i + 1, 2, number.amount)
    workbook.create_sheet("worksheet 2")
    sheet = workbook["worksheet 2"]
    for i, summon in enumerate(chain(targets, numbers)):
        sheet.cell(i + 1, 1, summon.account)
        sheet.cell(i + 1, 2, summon.amount)
    with BytesIO() as file:
        workbook.save(file)
        data_loader = ExcelDataLoader()
        data_loader.load(file)
        assert targets == data_loader.targets
        assert numbers == data_loader.numbers
        assert data_loader._loaded
