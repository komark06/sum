from io import BytesIO
from itertools import zip_longest

import openpyxl

from src.data_loader import Summons
from src.executor import BruteForceExecutor
from src.output import output_excel
from test.utils import FakeDataLoader


def test_output_excel():
    """Verify that output_excel successfully save output to file."""
    data_loader = FakeDataLoader()
    data_loader.load()
    results = BruteForceExecutor().calculate_all(
        data_loader.targets, data_loader.numbers
    )
    results = 2 * results  # Test multiple results.
    with BytesIO() as file:
        output_excel(results, data_loader, file)
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[workbook.sheetnames[0]]
        for row, target, number in zip_longest(
            sheet.iter_rows(min_row=2, values_only=True),
            data_loader.targets,
            data_loader.numbers,
        ):
            if target:
                assert target == Summons(row[0], target.date, row[1])
            if number:
                assert number == Summons(row[3], number.date, row[4])
        sheet = workbook[workbook.sheetnames[1]]
        for idx, result in enumerate(results):
            # Dynamically calculate the starting column for this result
            start_column = idx * 5 + 1
            assert sheet.cell(2, start_column).value == result.target.account
            assert (
                sheet.cell(2, start_column + 1).value == result.target.amount
            )
            for i, number in enumerate(result.subset, start=2):
                assert (
                    sheet.cell(
                        i,
                        start_column + 2,
                    ).value
                    == number.account
                )
                assert (
                    sheet.cell(i, start_column + 3, number.amount).value
                    == number.amount
                )
