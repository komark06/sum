import openpyxl
from openpyxl.utils import get_column_letter

from src.data_loader import AbstractDataLoader
from src.executor import Result


def output_excel(
    results: list[Result],
    data_loader: AbstractDataLoader,
    filename: str = "ex.xlsx",
):
    account_length = 29.0
    amount_length = 10.0
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"] = sheet["D1"] = "憑證號碼"
    sheet["B1"] = sheet["E1"] = "金額"
    for i, target in enumerate(data_loader.targets):
        sheet.cell(i + 2, 1, target.account)
        sheet.cell(i + 2, 2, target.amount)
    for i, number in enumerate(data_loader.numbers):
        sheet.cell(i + 2, 4, number.account)
        sheet.cell(i + 2, 5, number.amount)
    sheet.column_dimensions["A"].width = account_length
    sheet.column_dimensions["D"].width = account_length
    sheet.column_dimensions["B"].width = amount_length
    sheet.column_dimensions["E"].width = amount_length
    # Output
    output_account_length = 29.0
    output_amount_length = 10.0
    sheet_name = "配對表"
    wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    for idx, result in enumerate(results):
        # Dynamically calculate the starting column for this result
        start_column = idx * 5 + 1

        # Define column widths and headers
        column_widths = [
            output_account_length,
            output_amount_length,
            output_account_length,
            output_amount_length,
        ]
        headers = ["憑證號碼", "目標值", "憑證號碼", "配對值"]

        # Set column widths and headers
        for i, (width, header) in enumerate(
            zip(column_widths, headers), start=start_column
        ):
            sheet.column_dimensions[get_column_letter(i)].width = width
            sheet.cell(1, i, header)

        # Populate target account and amount
        sheet.cell(2, start_column, result.target.account)
        sheet.cell(2, start_column + 1, result.target.amount)

        # Populate subset data
        for i, number in enumerate(result.subset, start=2):
            sheet.cell(i, start_column + 2, number.account)
            sheet.cell(i, start_column + 3, number.amount)

    wb.save(filename)
