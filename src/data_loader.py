"""This module contains data loader that use to load data."""

import datetime
from abc import ABC, abstractmethod
from dataclasses import dataclass

import openpyxl


@dataclass
class Summons:
    """
    An Accounting summons.

    Attributes:
        account: The account of summon.
        date: The date of summon.
        amount: The amount of summon.
    """

    account: str
    date: datetime.date
    amount: int


class AbstractDataLoader(ABC):
    """
    An abstract data loader.

    Attributes:
        targets: The list of targets that we want to
            solve as subset sum target.
        numbers: The subset that we search for the sum
            of its subset is equal to target.
        _loaded: The flag indicate that data is loaded or not.
    """

    targets: list[Summons]
    numbers: list[Summons]
    _loaded: bool

    def __init__(self):
        """Initialize flag."""
        self._loaded = False

    @property
    def loaded(self):
        """
        Check data is loaded or not.

        If data is loaded, return True. Else, return false.
        """
        return self._loaded

    @abstractmethod
    def load(self):
        """Load data."""
        pass

    def sort(self):
        """Sort numbers and targets"""
        if self.loaded:
            self.numbers.sort(key=lambda x: (x.date, x.amount))
            self.targets.sort(key=lambda x: (x.date, x.amount))


class ExcelDataLoader(AbstractDataLoader):
    """A Data Loader load data from Excel."""

    def load(self, filename: str, reload=False):
        """Load data from Excel.

        Parameters:
            filename: The file path of excel.
            reload: The flag to decide reload or not.
        """
        if self.loaded and not reload:
            return
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[workbook.sheetnames[0]]
        targets = [
            row[0] for row in sheet.iter_rows(values_only=True) if row[0]
        ]
        self.targets = []
        self.numbers = []
        sheet = workbook[workbook.sheetnames[1]]
        for row in sheet.iter_rows(values_only=True):
            tag = row[0]
            date = datetime.date.fromisoformat(tag.split("-")[0])
            amount = abs(row[1])
            obj = Summons(tag, date, amount)
            if amount in targets:
                self.targets.append(obj)
                targets.remove(amount)
            else:
                self.numbers.append(obj)
        self._loaded = True
        self.sort()
